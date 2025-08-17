import datetime
import json
import logging
import os
import re
import traceback
from functools import wraps
from importlib import import_module
from io import BytesIO

from django.apps import apps
from django.conf import settings
from django.core.exceptions import FieldDoesNotExist
from django.core.mail import EmailMessage
from django.db import models
from django.db.models import Q
from django.http import HttpResponse
from drf_yasg.utils import swagger_auto_schema
from headquarters.models import RegionalHeadquarter, RegionalHeadquarterEmail
from openpyxl import Workbook
from pdfrw import PageMerge, PdfReader, PdfWriter
from regional_competitions.constants import MASS_REPORT_NUMBERS, MEDIA_PATH
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import HRFlowable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework import serializers, status

logger = logging.getLogger('regional_tasks')


def get_current_year():
    return datetime.datetime.now().year


def get_last_rcompetition_id():
    from .models import RCompetition
    last_r_competition = RCompetition.objects.last()
    return last_r_competition.id if last_r_competition else None


def regional_comp_regulations_files_path(instance, filename) -> str:
    """Функция для формирования пути сохранения файлов конкурса РШ.

    Сначала проверяет наличие атрибута `regional_headquarter`.
    Если атрибут отсутствует, ищет атрибут, начинающийся на `regional_r`,
    и через него обращается к `regional_headquarter`.

    :param instance: Экземпляр модели.
    :param filename: Имя файла. Добавляем к имени текущую дату и время.
    :return: Путь к изображению.
    """
    filename_parts = filename.rsplit('.', 1)
    base_filename = filename_parts[0][:25]
    file_extension = filename_parts[1] if len(filename_parts) > 1 else ''

    if hasattr(instance, 'regional_headquarter'):
        regional_hq_id = instance.regional_headquarter.id
    else:
        regional_r_attr = next(
            (getattr(instance, attr) for attr in dir(instance) if attr.startswith('regional_r') and
                hasattr(getattr(instance, attr), 'regional_headquarter')),
            None
        )
        if regional_r_attr:
            regional_hq_id = regional_r_attr.regional_headquarter.id
        else:
            raise AttributeError(
                "Не удалось найти атрибут regional_headquarter или атрибут, начинающийся с 'regional_r'."
            )

    return f'regional_comp/regulations/{instance}/{regional_hq_id}/{base_filename}.{file_extension}'


def swagger_schema_for_retrieve_method(serializer_cls):
    """Создает декоратор для метода retrieve, генерирующий Swagger схему."""

    def decorator(func):
        @swagger_auto_schema(responses={status.HTTP_200_OK: serializer_cls})
        @wraps(func)
        def wrapped(self, *args, **kwargs):
            return func(self, *args, **kwargs)
        return wrapped
    return decorator


def swagger_schema_for_district_review(serializer_cls):
    def decorator(func):
        @swagger_auto_schema(methods=['PUT'], request_body=serializer_cls)
        @wraps(func)
        def wrapped(*args, **kwargs):
            return func(*args, **kwargs)

        return wrapped

    return decorator


def swagger_schema_for_central_review(serializer_cls):
    def decorator(func):
        @swagger_auto_schema(methods=['PUT', 'DELETE'], request_body=serializer_cls)
        @wraps(func)
        def wrapped(*args, **kwargs):
            return func(*args, **kwargs)

        return wrapped

    return decorator


def swagger_schema_for_create_and_update_methods(serializer_cls):
    def decorator(func):
        @swagger_auto_schema(request_body=serializer_cls)
        @wraps(func)
        def wrapped(self, *args, **kwargs):
            return func(self, *args, **kwargs)

        return wrapped

    return decorator


def log_exception(func):
    @wraps(func)
    def wrapped(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            trb = traceback.format_exc()
            logger.exception(f'Возник Exception!!!: {e}\n{trb}')

    return wrapped


def get_report_number_by_class_name(link):
    """
    Получает номер отчета для классов с названием,
    соответствующего шаблону `RegionalR<номер_отчета>` или `RegionalReport<номер_отчета>`.
    """
    NUMBER_INDEX = 1
    class_name = link.__class__.__name__
    pattern = r'Regional(?:Report)?(\d+)'
    match_regional_report = re.search(pattern, class_name)

    if class_name == 'str':
        link_length = len(link)
        if link_length >= 13 and link[12].isdigit():
            return link[9:13]
        if link_length >= 12 and link[11].isdigit():
            return link[9:12]
        if link_length >= 11 and link[10].isdigit():
            return link[9:11]
        return link[9]

    if match_regional_report:
        return match_regional_report.group(NUMBER_INDEX)

    pattern2 = r'RegionalR(\d+)'
    match_regional_r = re.search(pattern2, class_name)

    if match_regional_r:
        return match_regional_r.group(NUMBER_INDEX)

    return None


def get_emails(report_instance) -> list:
    if settings.DEBUG:
        return settings.TEST_EMAIL_ADDRESSES
    addresses = []

    if isinstance(report_instance, RegionalHeadquarter):
        regional_headquarter = report_instance
    elif isinstance(report_instance, int):
        regional_headquarter = RegionalHeadquarter.objects.get(pk=report_instance)
    else:
        regional_headquarter = report_instance.regional_headquarter

    try:
        rhq_emails = RegionalHeadquarterEmail.objects.filter(regional_headquarter=regional_headquarter)
        for rhq_email in rhq_emails:
            addresses.append(
                rhq_email.email
            )
        if regional_headquarter.id == 74 and settings.PRODUCTION and not settings.DEBUG:
            addresses.append("rso.71@yandex.ru")
    except RegionalHeadquarterEmail.DoesNotExist:
        logger.warning(
            f'Не найден почтовый адрес в RegionalHeadquarterEmail '
            f'для РШ {regional_headquarter} ID {regional_headquarter.id}'
        )
    addresses.append('rso.login@yandex.ru')
    addresses.append('delightxxls@gmail.com')
    addresses.append('olegfreon@yandex.ru')
    return addresses


@log_exception
def send_email_with_attachment(
        subject: str, message: str, recipients: list, file_path: str, additional_file_path: str = None
        ):
    mail = EmailMessage(
        subject=subject,
        body=message,
        from_email=settings.EMAIL_HOST_USER,
        to=recipients
    )
    if file_path:
        with open(file_path, 'rb') as f:
            mail.attach(file_path.split('/')[-1], f.read(), 'application/octet-stream')
    if additional_file_path:
        with open(additional_file_path, 'rb') as f:
            mail.attach(additional_file_path.split('/')[-1], f.read(), 'application/octet-stream')
    mail.send()


def generate_pdf_report_part_1(report_id) -> str:
    from regional_competitions.models import StatisticalRegionalReport

    try:
        report = StatisticalRegionalReport.objects.get(pk=report_id)
    except:
        raise Exception(f'В generate_pdf_report_part_1 передан некорректный id первого отчета: {report_id}')

    pdf_file_name = f"Отчет_ч1_РСО_{report.regional_headquarter}.pdf"
    pdf_file_path = os.path.join(settings.MEDIA_ROOT, pdf_file_name)

    pdfmetrics.registerFont(TTFont(
        'Times_New_Roman',
        os.path.join(
            str(settings.BASE_DIR),
            'templates',
            'samples',
            'fonts',
            'times.ttf'
        )
    ))

    temp_pdf_file_path = os.path.join(settings.MEDIA_ROOT, f"Temp_{pdf_file_name}")
    doc = SimpleDocTemplate(temp_pdf_file_path, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=30,
                            bottomMargin=18)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name='CustomTitle', fontName='Times_New_Roman', fontSize=18, spaceAfter=20, alignment=1, leading=24
        )
    )
    styles.add(
        ParagraphStyle(name='Times_New_Roman', fontName='Times_New_Roman', fontSize=12)
    )

    elements.append(Spacer(1, 90))
    elements.append(
        Paragraph(
            f'Отчет о деятельности регионального отделения за 2024 год. Часть 1. {report.regional_headquarter}',
            styles['CustomTitle']
        )
    )
    elements.append(Spacer(1, 15))
    data = []

    for field in report._meta.fields:
        if field.name == 'id':
            continue
        field_name = field.verbose_name
        field_value = getattr(report, field.name, '')
        if isinstance(field_value, (int, float, str)):
            field_value = str(field_value)
        elif field_value is None:
            field_value = ''
        else:
            field_value = str(field_value)

        data.append([Paragraph(f"<b>{field_name}:</b>", styles['Times_New_Roman']), Paragraph(field_value, styles['Times_New_Roman'])])

    if data:
        table = Table(data, colWidths=[5 * cm, 10 * cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6699CC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#003366')),
            ('FONTNAME', (0, 0), (-1, -1), 'Times_New_Roman'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 15))

    elements.append(PageBreak())

    elements.append(
        Paragraph("Дополнительные данные:", styles['CustomTitle'])
    )
    elements.append(Spacer(1, 10))

    additional_data = []

    for nested_dict in report.additional_statistics.all():
        additional_data.append(
            [
                Paragraph(nested_dict.name, styles['Times_New_Roman']), 
                Paragraph(str(nested_dict.value), styles['Times_New_Roman'])
            ]
        )

    if additional_data:
        additional_table = Table(additional_data, colWidths=[5 * cm, 10 * cm])
        additional_table.setStyle(TableStyle([
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#003366')),  # Same text color for all rows
            ('FONTNAME', (0, 0), (-1, -1), 'Times_New_Roman'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),  # Same background color for all rows
        ]))
        elements.append(additional_table)

    doc.build(elements)

    template_pdf_path = os.path.join(
        str(settings.BASE_DIR),
        'templates',
        'samples',
        'header_regional_r.pdf'
    )

    template_reader = PdfReader(template_pdf_path)
    content_reader = PdfReader(temp_pdf_file_path)

    writer = PdfWriter()
    for i, content_page in enumerate(content_reader.pages):
        if i == 0:
            template_page = template_reader.pages[0]
            PageMerge(template_page).add(content_page, prepend=False).render()
            writer.addpage(template_page)
        else:
            writer.addpage(content_page)

    writer.write(pdf_file_path)
    os.remove(temp_pdf_file_path)
    return pdf_file_path


def get_verbose_names_and_values(serializer, full_path: bool = False) -> dict:
    """Возвращает словарь с названиями полей и значениями полей из сериализатора."""

    custom_verbose_names_dict = {
        'events': 'Мероприятия/Проекты',
        'links': 'Ссылки',
    }
    custom_values_dict = {
        'True': 'Да',
        'False': 'Нет',
        'None': '-',
    }
    verbose_names_and_values = {}
    model_meta = serializer.Meta.model._meta
    instance = serializer.instance

    for field_name, field in serializer.fields.items():
        field_value = getattr(instance, field_name, None)

        if str(field_value) in custom_values_dict:
            field_value = custom_values_dict[str(field_value)]

        if str(field_value) in custom_values_dict:
            field_value = custom_values_dict[str(field_value)]

        if isinstance(field, serializers.ListSerializer):
            if hasattr(field_value, 'all') and callable(field_value.all):
                nested_verbose_names_and_values = []
                for nested_instance in field_value.all():
                    nested_serializer = field.child.__class__(nested_instance)
                    nested_verbose_names_and_values.append(get_verbose_names_and_values(nested_serializer, full_path))
                verbose_names_and_values[field_name] = nested_verbose_names_and_values
            else:
                verbose_names_and_values[field_name] = field_value

        elif isinstance(field, serializers.ModelSerializer):
            if field_value is not None and hasattr(field_value, '_meta'):
                nested_serializer = field.__class__(field_value)
                nested_verbose_names_and_values = get_verbose_names_and_values(nested_serializer, full_path)
                for nested_field_name, nested_verbose_name_and_value in nested_verbose_names_and_values.items():
                    verbose_names_and_values[f"{field_name}.{nested_field_name}"] = nested_verbose_name_and_value
            else:
                verbose_names_and_values[field_name] = field_value

        else:
            try:
                verbose_name = model_meta.get_field(field_name).verbose_name
                if hasattr(field_value, '__str__'):
                    field_value = str(field_value)
                    if not field_value.startswith('http') and full_path is False:
                        field_value = os.path.basename(field_value)
                    if field_value.startswith('regional_comp') and full_path is True:
                        field_value = MEDIA_PATH + field_value

                verbose_names_and_values[field_name] = (verbose_name, field_value)
            except FieldDoesNotExist:
                pass
            except AttributeError:
                verbose_names_and_values[field_name] = (
                    custom_verbose_names_dict.get(str(field_name).lower(), field_name), field_value
                )

    return verbose_names_and_values


def get_headers_values(fields_dict: dict, prefix: str = '') -> dict:
    """Формирует плоский словарь для заголовков и значений листа Excel."""

    flat_dict = {}

    for value in fields_dict.values():
        if isinstance(value, list):
            for item in value:
                try:
                    nested_prefix = f'{prefix}{item["id"][0]}_{item["id"][1]}.'
                except KeyError:
                    nested_prefix = prefix
                nested_dict = get_headers_values(fields_dict=item, prefix=nested_prefix)
                flat_dict.update(nested_dict)
        elif isinstance(value, tuple):
            flat_dict[f'{prefix}{value[0]}'] = value[1]

    return flat_dict


def get_model_and_serializer(report_number: str):
    """Возвращает модель и класс сериализатора для заданного номера отчета."""
    model_name = 'RegionalR' + report_number
    try:
        model = apps.get_model('regional_competitions_2025', model_name)
    except LookupError:
        model = None

    if not model:
        raise ValueError(f'Модель {model_name} не найдена.')

    serializer_name = 'RegionalReport' + report_number + 'Serializer'
    serializers_module = import_module('regional_competitions_2025.serializers')
    serializer_class = getattr(serializers_module, serializer_name, None)

    if not serializer_class:
        raise ValueError(f'Сериализатор {serializer_name} не найден.')

    return model, serializer_class


def create_excel_file(title: str, headers: list, rows: list) -> BytesIO:
    """Создает Excel-файл с указанными заголовками и строками."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title

    if headers:
        worksheet.append(headers)

    for row in rows:
        worksheet.append(row)

    file_content = BytesIO()
    workbook.save(file_content)
    file_content.seek(0)
    return file_content


def get_report_data(model, serializer_class):
    """Получает данные из модели и сериализует их."""
    if model.__name__ in ['RegionalR2', 'RegionalR3', 'RegionalR7', 'RegionalR8', 'RegionalR14', 'RegionalR15']:
        reports = model.objects.all()
    else:
        reports = model.objects.filter(
            Q(verified_by_chq=True) | Q(verified_by_chq__isnull=True),
        )
    rows = []
    headers_written = False

    for report in reports:
        serializer = serializer_class(report)
        flat_data_dict = get_headers_values(
            fields_dict=get_verbose_names_and_values(serializer, full_path=True)
        )

        if not headers_written:
            headers = list(flat_data_dict.keys())
            headers_written = True

        rows.append(list(flat_data_dict.values()))

    return headers if headers_written else [], rows


def generate_report_response(title: str, file_content: BytesIO) -> HttpResponse:
    """Создает HTTP-ответ с Excel-файлом."""
    response = HttpResponse(
        file_content.read(),
        content_type=(
            'application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.sheet'
        )
    )
    response['Content-Disposition'] = f'attachment; filename={title}.xlsx'
    return response


def get_all_reports_from_competition(report_number: int):
    """Возвращает xlsx со всеми отчетами для заданного показателя."""
    model, serializer_class = get_model_and_serializer(str(report_number))
    headers, rows = get_report_data(model, serializer_class)
    file_content = create_excel_file(f'Reports_{report_number}', headers, rows)
    return generate_report_response(f'Reports_{report_number}', file_content)


def get_reports_from_mass_competitions(main_report_number: int):
    """Возвращает xlsx с массовыми показателями на одном листе."""
    all_headers = None
    all_rows = []
    entries_dict = {
        6: 115,
        9: 11,
        10: 2
    }
    entries_number = entries_dict.get(main_report_number, 0)
    for sub_number in range(1, entries_number + 1):
        report_number = f"{main_report_number}{sub_number}"
        try:
            model, serializer_class = get_model_and_serializer(report_number)
            headers, rows = get_report_data(model, serializer_class)

            if not all_headers:
                all_headers = headers
            all_rows.extend(rows)
        except ValueError as e:
            logger.error(f'Возник Exception!!!: {e}\n{traceback.format_exc()}', exc_info=True)
            continue

    if not all_headers:
        raise ValueError(f'Не удалось найти данные для показателя {main_report_number}.')

    file_content = create_excel_file(f'Reports_{main_report_number}', all_headers, all_rows)
    return generate_report_response(f'Reports_{main_report_number}', file_content)


def get_all_models(module_name: str):
    """Возвращает список всех моделей RegionalR из заданного модуля и динамически созданных моделей."""
    all_models = ['DumpStatisticalRegionalReport', 'StatisticalRegionalReport',]
    pattern = re.compile(r'^RegionalR\d+$')

    for model in apps.get_models():
        model_name = model.__name__
        if issubclass(model, models.Model) and pattern.match(model_name):
            all_models.append(model_name)

    return all_models


def generate_rhq_xlsx_report(regional_headquarter_id: int) -> HttpResponse:
    """Выгрузка отчётов определенного РШ в формате Excel."""

    models_list = get_all_models('regional_competitions.models')
    first_ws_is_filled = False
    workbook = Workbook()
    STATISTICAL_REPORT_POSITION = 1

    for model_name in models_list:
        report_number = model_name.split('RegionalR')[1]

        model = apps.get_model('regional_competitions', model_name)
        if model_name == 'DumpStatisticalRegionalReport':
            if not model.objects.filter(regional_headquarter_id=regional_headquarter_id).exists():
                continue
            else:
                models_list.pop(STATISTICAL_REPORT_POSITION)
        if report_number == '14':
            instance = model.objects.filter(report_12__regional_headquarter_id=regional_headquarter_id).first()
            if instance is None:
                continue
            worksheet = workbook.create_sheet(report_number)
            worksheet.append(['Очки'])
            worksheet.append([str(instance.score)])
            continue

        instance = model.objects.filter(regional_headquarter_id=regional_headquarter_id).first()
        if instance is None:
            continue

        serializer_name = model_name + 'Serializer'
        serializers_module = import_module('regional_competitions.serializers')
        serializer_class = getattr(serializers_module, serializer_name, None)
        if serializer_class is None:
            continue

        serializer_data = serializer_class(instance)
        report_data = get_headers_values(
            get_verbose_names_and_values(
                serializer_data,
                full_path=True
            )
        )
        if model_name == 'DumpStatisticalRegionalReport' or model_name == 'StatisticalRegionalReport':
            sheet_name = 'Статистический отчет'
        elif report_number == '101' or report_number == '102':
            sheet_name = '10'
        else:
            sheet_name = report_number[0] if report_number[0] in MASS_REPORT_NUMBERS else report_number

        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            if not first_ws_is_filled:
                worksheet = workbook.active
                worksheet.title = sheet_name
                first_ws_is_filled = True
            else:
                worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(list(report_data.keys()))

        worksheet.append(list(report_data.values()))

    file_content = BytesIO()
    workbook.save(file_content)
    file_content.seek(0)
    response = HttpResponse(
        file_content.read(),
        content_type=(
            'application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.sheet'
        )
    )
    response['Content-Disposition'] = 'attachment; filename=RO_report.xlsx'
    return response


def generate_pdf_report_part_2(regional_headquarter_id: int) -> str:
    """Generate an improved PDF report for Part 2."""
    from regional_competitions.serializers import REPORTS_SERIALIZERS

    try:
        regional_hq = RegionalHeadquarter.objects.get(id=regional_headquarter_id).name
    except RegionalHeadquarter.DoesNotExist:
        regional_hq = ''

    pdf_file_name = f"Отчет_ч2_РСО_{regional_hq}.pdf"
    pdf_file_path = os.path.join(settings.MEDIA_ROOT, pdf_file_name)

    pdfmetrics.registerFont(TTFont(
        'Times_New_Roman',
        os.path.join(
            str(settings.BASE_DIR),
            'templates',
            'samples',
            'fonts',
            'times.ttf'
        )
    ))

    doc_style = getSampleStyleSheet()

    doc_style['Normal'].fontName = 'Times_New_Roman'
    doc_style['Normal'].fontSize = 10
    doc_style['Normal'].leading = 10

    doc_style.add(
        ParagraphStyle(
            name='CustomTitle',
            parent=doc_style['Normal'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_CENTER,
            leading=24,
            textColor=colors.HexColor('#003366')
        )
    )
    doc_style.add(
        ParagraphStyle(
            name='NestedField',
            parent=doc_style['Normal'],
            leftIndent=20,
            spaceAfter=10,
        )
    )
    doc_style.add(
        ParagraphStyle(
            name='SectionHeader',
            parent=doc_style['Normal'],
            fontSize=12,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.HexColor('#003366')
        )
    )

    temp_pdf_file_path = os.path.join(settings.MEDIA_ROOT, f"Temp_{pdf_file_name}")
    doc = SimpleDocTemplate(
        temp_pdf_file_path, pagesize=A4,
        rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=18
    )
    elements = []

    elements.append(Spacer(1, 90))
    elements.append(
        Paragraph(
            f'<b>Отчет о деятельности регионального отделения за 2024 год. Часть 2. {regional_hq}</b>',
            doc_style['CustomTitle']
        )
    )
    elements.append(Spacer(1, 15))

    first_block = True

    for serializer_class in REPORTS_SERIALIZERS:
        queryset = serializer_class.Meta.model.objects.filter(regional_headquarter_id=regional_headquarter_id)
        if not queryset.exists():
            continue
        for instance in queryset:
            serializer = serializer_class(instance)
            verbose_names_and_values = get_verbose_names_and_values(serializer)

            if first_block:
                elements.append(
                    HRFlowable(
                        width="100%", thickness=2, color=colors.HexColor('#003366'),
                        spaceBefore=10, spaceAfter=10
                    )
                )
                first_block = False

            elements.append(Paragraph(f'<b>{serializer.Meta.model._meta.verbose_name}</b>', doc_style['SectionHeader']))
            elements.append(Spacer(1, 10))

            add_verbose_names_and_values_to_pdf(verbose_names_and_values, elements, doc_style)

            elements.append(
                HRFlowable(
                    width="100%", thickness=2, color=colors.HexColor('#003366'),
                    spaceBefore=10, spaceAfter=20
                )
            )

    doc.build(elements)

    template_pdf_path = os.path.join(
        str(settings.BASE_DIR),
        'templates',
        'samples',
        'header_regional_r.pdf'
    )
    template_reader = PdfReader(template_pdf_path)
    content_reader = PdfReader(temp_pdf_file_path)

    writer = PdfWriter()
    for i, content_page in enumerate(content_reader.pages):
        if i == 0:
            template_page = template_reader.pages[0]
            PageMerge(template_page).add(content_page, prepend=False).render()
            writer.addpage(template_page)
        else:
            writer.addpage(content_page)

    writer.write(pdf_file_path)
    os.remove(temp_pdf_file_path)
    return pdf_file_path


def add_verbose_names_and_values_to_pdf(
    verbose_names_and_values: dict,
    elements: list,
    styles,
    indent=0,
    is_nested=False
):
    """Adds fields and values to the PDF with improved design for nested structures."""

    excluded_fields = [
        'id',
        'regional_headquarter',
        'regional_r',
        'verified_by_chq',
        'verified_by_dhq',
        'score',
        'created_at',
        'updated_at',
        'file_size',
        'file_type',
        'regional_version',
        'district_version',
        'central_version',
        'rejecting_reasons'
    ]

    primary_color = colors.HexColor('#003366')
    header_background_color = colors.HexColor('#F0F0F0')
    cell_background_color = colors.white

    nested_structures = []

    for field_name, field_value in verbose_names_and_values.items():
        if any(field_name.startswith(excluded_field) for excluded_field in excluded_fields):
            continue

        if isinstance(field_value, list):
            nested_structures.append((field_name, field_value))
            continue

        if isinstance(field_value, tuple) and len(field_value) == 2:
            verbose_name, field_value_content = field_value
        else:
            continue

        if not isinstance(field_value_content, dict):
            # Directly add the content as Paragraphs
            elements.append(Paragraph(f"<b>{verbose_name}:</b> {str(field_value_content)}", styles['Normal']))
            elements.append(Spacer(1, 5))

    for nested_name, nested_items in nested_structures:
        singular_names = {
            "projects": "Проект",
            "events": "Мероприятие",
            "links": "Ссылка",
        }
        singular_name = singular_names.get(nested_name, nested_name)

        for idx, item in enumerate(nested_items, start=1):
            if isinstance(item, dict):
                elements.append(Paragraph(f"<b>{singular_name} {idx}</b>", styles['NestedField']))
                add_verbose_names_and_values_to_pdf(item, elements, styles, indent + 1, is_nested=True)
            elements.append(Spacer(1, 5))


def get_participants(report, model):
    """Возвращает количество участников с уплаченными взносами из первого показателя."""

    ro_id = report.regional_headquarter.id

    try:
        participants = model.objects.filter(
            verified_by_chq=True,
            regional_headquarter_id=ro_id
        ).first().score
    except AttributeError:
        return
    return participants


def return_comp_from_logs(regional_headquarter_id, r_number, r_max_subnumber, log_model):
    """

    Функция записывает ссылки в связанные таблицы ссылок к показателям РО.
    Данные берем из логов RVerificationLog.
    """


    for index in range(1, r_max_subnumber+1):
        report_number = int(r_number + str(index))
        try:
            report_data = log_model.objects.filter(
                regional_headquarter=regional_headquarter_id,
                is_regional_data=True,
                report_number=report_number,
            ).last().data
            if isinstance(report_data, str):
                report_data = json.loads(report_data)
            links = report_data.get('links', None)
            if links:
                model_name = 'RegionalR' + str(report_number)
                model = apps.get_model('regional_competitions', model_name)
                instance = model.objects.filter(
                    regional_headquarter=regional_headquarter_id,
                    verified_by_dhq=True
                ).first()
                for item in links:
                    link = item.get('link', None)
                    if link:
                        instance.links.create(
                            link=link,
                        )
        except log_model.DoesNotExist:
            continue
        except AttributeError:
            continue
        except Exception as e:
            logger.exception(f'Исключение при получении ссылок из логов RVerificationLog: {e}')


def get_r_competition_by_year(year, r_model):
    try:
        year = int(year)
        r_competition = r_model.objects.get(year=year)
    except (ValueError, r_model.DoesNotExist):
        from regional_competitions_2025.utils import get_current_year
        r_competition = r_model.objects.get(year=get_current_year())
    else:
        from regional_competitions_2025.utils import get_current_year
        r_competition = r_model.objects.get(year=get_current_year())
    return r_competition
