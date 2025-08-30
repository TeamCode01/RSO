import logging
from io import BytesIO

from celery import shared_task
from openpyxl import Workbook
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from urllib.parse import unquote
from openpyxl.styles import Alignment

from reports.utils import (
    get_attributes_of_uniform_data, get_commander_school_data, get_detachment_q_results, get_membership_fee_data, get_regional_ranking_results,
    get_regions_users_data, get_safety_results,
    get_competition_participants_contact_data,
    get_competition_participants_data,
    get_q5_data, get_q6_data, get_q7_data, get_q8_data, get_q9_data, get_q10_data, get_q11_data, get_q12_data, get_q13_data, get_q14_data,
    get_q15_data, get_q16_data, get_q17_data,get_q18_data, get_q19_data, get_district_hq_data,
    get_regional_hq_data, get_detachment_data, get_educational_hq_data, get_local_hq_data, get_central_hq_data,
    get_direction_data, get_users_registry_data, get_template_data
)


logger = logging.getLogger('tasks')


@shared_task
def generate_excel_file(headers, worksheet_title, filename, data_func, fields=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheet_title
    worksheet.append(headers)
    
    if fields and 'section_headers' in fields:
        print(f"Processing section_headers: {fields['section_headers']}")
        for section in fields['section_headers']:
            text = section['text']
            merge_cells = section['merge_cells']
            worksheet.cell(row=merge_cells[0], column=merge_cells[1]).value = text
            worksheet.merge_cells(
                start_row=merge_cells[0],
                start_column=merge_cells[1],
                end_row=merge_cells[2],
                end_column=merge_cells[3]
            )
            worksheet.cell(row=merge_cells[0], column=merge_cells[1]).alignment = Alignment(horizontal='center')

    headers_row = fields.get('headers_row', 1) if fields else 1
    
    print(f"Headers: {headers}")
    for col_idx, header in enumerate(headers, 1):
        worksheet.cell(row=headers_row, column=col_idx).value = header

    def default_process_row(index, row):
        return [index] + list(row)

    data = None
    print('зашли в таску')
    match data_func:
        case 'safety_test_results':
            data = get_safety_results()
        case 'detachment_q_results':
            data = get_detachment_q_results(settings.COMPETITION_ID)
        case 'contact_data':
            data = get_competition_participants_contact_data()
        case 'competition_participants':
            data = get_competition_participants_data()
        case 'commander_school':
            data = get_commander_school_data(
                competition_id=settings.COMPETITION_ID)
        case 'regions_users_data':
            data = get_regions_users_data()
        case 'get_q5_data':
            data = get_q5_data(settings.COMPETITION_ID)
        case 'get_q6_data':
            data = get_q6_data(settings.COMPETITION_ID)
        case 'get_q7_data':
            data = get_q7_data(settings.COMPETITION_ID)
        case 'get_q8_data':
            data = get_q8_data(settings.COMPETITION_ID)
        case 'get_q9_data':
            data = get_q9_data(settings.COMPETITION_ID)
        case 'get_q10_data':
            data = get_q10_data(settings.COMPETITION_ID)
        case 'get_q11_data':
            data = get_q11_data(settings.COMPETITION_ID)
        case 'get_q12_data':
            data = get_q12_data(settings.COMPETITION_ID)
        case 'get_q13_data':
            data = get_q13_data(settings.COMPETITION_ID)
        case 'get_q14_data':
            data = get_q14_data(settings.COMPETITION_ID)
        case 'get_q15_data':
            data = get_q15_data(settings.COMPETITION_ID)
        case 'get_q16_data':
            data = get_q16_data(settings.COMPETITION_ID)
        case 'get_q17_data':
            data = get_q17_data(settings.COMPETITION_ID)
        case 'get_q18_data':
            data = get_q18_data(settings.COMPETITION_ID)
        case 'get_q19_data':
            data = get_q19_data(settings.COMPETITION_ID)
        case 'membership_fee':
            data = get_membership_fee_data(
                competition_id=settings.COMPETITION_ID)
        case 'attributes_of_uniform':
            data = get_attributes_of_uniform_data(
                competition_id=settings.COMPETITION_ID)
        case 'get_central_hq_data':
            data = get_central_hq_data(fields)
        case 'get_district_hq_data':
            data = get_district_hq_data(fields)
        case'get_regional_hq_data':
            data = get_regional_hq_data(fields)
        case 'get_educational_hq_data':
            data = get_educational_hq_data(fields)
        case 'get_local_hq_data':
            data = get_local_hq_data(fields)
        case 'get_detachment_data':
            data = get_detachment_data(fields)
        case 'get_direction_data':
            data = get_direction_data(fields)
        case 'get_users_registry_data':
            data = get_users_registry_data(fields)
        case 'get_regional_ranking':
            print('смэтчились')
            data = get_regional_ranking_results()
        case 'get_template_data':
            data = get_template_data()
        case _:
            logger.error(f"Неизвестное значение data_func: {data_func}")
            return
    if not data:
        logger.warning(
            'Вызов функции не соответствующей кейсу для вызова функции с data')
        return
    # logger.info('Формируем файл с данными')
    for index, row in enumerate(data, start=1):
        processed_row = default_process_row(index, row)
        worksheet.append(processed_row)

    file_content = BytesIO()
    workbook.save(file_content)
    file_content.seek(0)
    decoded_filename = unquote(filename)

    content = ContentFile(file_content.read())
    file_path = default_storage.save(f'to_delete_content/{decoded_filename}', content)
    # logger.info('Возврат эксель файла')
    return file_path


@shared_task
def delete_temp_reports_task():
    for file_name in default_storage.listdir('to_delete_content')[1]:
        default_storage.delete(f'to_delete_content/{file_name}')
